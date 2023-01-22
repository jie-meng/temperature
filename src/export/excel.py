# coding: utf-8

from openpyxl import Workbook
from typing import List
from src.models.data import Data


def write_output(data_array: List[Data], xlsx_file: str):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = '总体数据分析'
    ws1['A1'] = '图片'
    ws1['B1'] = '是否正常'
    ws1['C1'] = '点数'
    ws1['D1'] = '最小值'
    ws1['E1'] = '最大值'
    ws1['F1'] = '均值'
    ws1['G1'] = '方差'
    ws1['H1'] = '标准差'
    ws1['I1'] = '变异系数'

    ws2 = wb.create_sheet(title = '数据集')

    for idx, data in enumerate(data_array):
        ws1.cell(row = idx + 2, column = 1, value = data.image_name)
        ws1.cell(row = idx + 2, column = 2, value = data.is_ok)
        ws1.cell(row = idx + 2, column = 3, value = data.point_count)
        ws1.cell(row = idx + 2, column = 4, value = data.min_value)
        ws1.cell(row = idx + 2, column = 5, value = data.max_value)
        ws1.cell(row = idx + 2, column = 6, value = data.mean)
        ws1.cell(row = idx + 2, column = 7, value = data.var)
        ws1.cell(row = idx + 2, column = 8, value = data.std)
        ws1.cell(row = idx + 2, column = 9, value = data.coefficient_of_variation)

        ws2.cell(row = 1, column = idx + 1, value = data.image_name)
        for data_idx, value in enumerate(data.data_array):
            ws2.cell(row = data_idx + 2, column = idx + 1, value = value)

    wb.save(filename = xlsx_file)
