# -*- coding: utf-8 -*-

import os
import re
import numpy as np
import click
from dataclasses import dataclass
from typing import List, Dict
from pathlib import Path
from PIL import Image
from openpyxl import Workbook


@dataclass
class Data:
    image_name: str = ''
    is_ok: bool = False
    point_count: int = 0
    min_value: float = 0.0
    max_value: float = 0.0
    mean: float = 0.0
    var: float = 0.0
    std: float = 0.0
    coefficient_of_variation: float = 0.0
    data_array: List[float] = None


def raw_diff(color1: List[float], color2: List[float]) -> float:
    return abs(color1[0] - color2[0]) + abs(color1[1] - color2[1]) + abs(color1[2] - color2[2]) + abs(color1[3] - color2[3])


def load_ruler(root_path: str) -> List[List[float]]:
    ruler_image = Image.open(root_path + '/ruler.png')
    pix = ruler_image.load()

    ruler = []
    for i in range(ruler_image.size[1] - 1, -1, -1):
        ruler.append(pix[4, i])

    ruler_image.close()

    return ruler


def parse_config(config: str) -> Dict[str, str]:
    with open(config, 'r') as f:
        lines = list(filter(lambda y: re.match('\w+=-?\d+\.?\d*', y), map(lambda x: x.strip(), f.readlines())))

        config_data = {}
        for line in lines:
            result = line.split('=')
            config_data[result[0]] = result[1]

        return config_data


def calc_image_data(debug: bool, ruler: List[List[float]], image: str, config: str) -> Data:
    config_data = parse_config(config)
    ruler_min = float(config_data['ruler_min'])
    ruler_max = float(config_data['ruler_max'])


    target_im = Image.open(image)
    pix = target_im.load()

    print('Calculating {0} (size {1}x{2}) ...'.format(image, target_im.size[0], target_im.size[1]))

    all_value_array = []
    for i in range(0, target_im.size[0]):
        for j in range(0, target_im.size[1]):
            # Only calculate non-transparent pixels
            if pix[i, j][3] != 0:
                if debug:
                    print('Include>>> [{0},{1}] R:{2}, G:{3}, B:{4}, A:{5}'.format(i, j, pix[i, j][0], pix[i, j][1], pix[i, j][2], pix[i, j][3]))

                idx = ruler.index(min(ruler, key = lambda x: raw_diff(x, pix[i, j])))
                value = (idx / len(ruler)) * (ruler_max - ruler_min) + ruler_min
                all_value_array.append(value)
            else:
                if debug:
                    print('Exclude>>> [{0},{1}] R:{2}, G:{3}, B:{4}, A:{5}'.format(i, j, pix[i, j][0], pix[i, j][1], pix[i, j][2], pix[i, j][3]))


    
    if len(all_value_array) == target_im.size[0] * target_im.size[1]:
        print('Warning: this image does not cliped correctly!')

    target_im.close()

    data = Data(
        image_name = Path(image).stem,
        point_count = len(all_value_array),
        min_value = np.min(all_value_array),
        max_value = np.max(all_value_array),
        mean = np.mean(all_value_array),
        var = np.var(all_value_array),
        std = np.std(all_value_array),
        data_array = all_value_array,
    )

    data.coefficient_of_variation = data.std / data.mean
    data.is_ok = True

    return data


def write_output(data_array: List[Data], xlsx_file: str):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = '总体数据分析'
    ws1['A1'] = '图片'
    ws1['B1'] = '是否正常'
    ws1['C1'] = '点数'
    ws1['D1'] = '最小值'
    ws1['E1'] = '最大值'
    ws1['F1'] = '均值'
    ws1['G1'] = '方差'
    ws1['H1'] = '标准差'
    ws1['I1'] = '变异系数'

    ws2 = wb.create_sheet(title = '数据集')

    for idx, data in enumerate(data_array):
        ws1.cell(row = idx + 2, column = 1, value = data.image_name)
        ws1.cell(row = idx + 2, column = 2, value = data.is_ok)
        ws1.cell(row = idx + 2, column = 3, value = data.point_count)
        ws1.cell(row = idx + 2, column = 4, value = data.min_value)
        ws1.cell(row = idx + 2, column = 5, value = data.max_value)
        ws1.cell(row = idx + 2, column = 6, value = data.mean)
        ws1.cell(row = idx + 2, column = 7, value = data.var)
        ws1.cell(row = idx + 2, column = 8, value = data.std)
        ws1.cell(row = idx + 2, column = 9, value = data.coefficient_of_variation)

        ws2.cell(row = 1, column = idx + 1, value = data.image_name)
        for data_idx, value in enumerate(data.data_array):
            ws2.cell(row = data_idx + 2, column = idx + 1, value = value)

    wb.save(filename = xlsx_file)


def process_collection(debug: bool, ruler: List[List[float]], root_path: str, coll: str):
    coll_path = '{0}/images/{1}'.format(root_path, coll)
    print('\nProcessing collection {0} ...'.format(coll_path))

    # Find images
    images = list(filter(lambda x: x.endswith('.png'), os.listdir(coll_path)))
    images.sort()

    # Process each image
    data_array: List[Data] = []
    for image in images:
        image_path = '{0}/{1}'.format(coll_path, image)
        config_path = '{0}/{1}.txt'.format(coll_path, os.path.splitext(image)[0])

        if not os.path.isfile(config_path):
            print('Warning: cannot find {0}'.format(config_path))
            continue

        try:
            data_array.append(calc_image_data(debug, ruler, image_path, config_path))
        except BaseException:
            data_array.append(Data(image_name = Path(image).stem, is_ok = False, data_array = []))
            continue

    # Write xlsx
    xlsx_file = '{0}/{1}.xlsx'.format(coll_path, coll)
    print('Generate {0} ...'.format(xlsx_file))

    if len(data_array) > 0:
        write_output(data_array, xlsx_file)


@click.command()
@click.option('--debug', prompt = 'Debug mode?', default = 'n', required = True, type=click.Choice(['y', 'n']), help = 'Enable debug or not')
def cli(debug: str):
    """temperature-cli - Command line tool to analyze temperature image data"""
    # Find current script absolute path
    root_path = os.path.dirname(os.path.realpath(__file__))

    # Load ruler
    ruler = load_ruler(root_path)

    # Find collections
    collections = list(filter(lambda x: os.path.isdir('{0}/images/{1}'.format(root_path, x)), os.listdir('{0}/images'.format(root_path))))
    collections.sort()

    for coll in collections:
        process_collection(debug == 'y', ruler, root_path, coll)

    print('\nDone!')

if __name__ == '__main__':
    cli() 
    