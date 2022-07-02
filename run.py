# -*- coding: utf-8 -*-

import os
import re
from dataclasses import dataclass
from typing import List, Dict
from pathlib import Path
from PIL import Image
from openpyxl import Workbook
import numpy as np


@dataclass
class Data:
    image_name: str = ''
    point_count: int = 0
    min_value: float = 0.0
    max_value: float = 0.0
    mean: float = 0.0
    var: float = 0.0
    std: float = 0.0
    coefficient_of_variation: float = 0.0
    data_array: List[float] = None


def raw_diff(color1: List[float], color2: List[float]) -> float:
    return abs(color1[0] - color2[0]) + abs(color1[1] - color2[1]) + abs(color1[2] - color2[2]) + abs(color1[3] - color2[3])


def load_ruler(root_path: str) -> List[float]:
    ruler_image = Image.open(root_path + '/ruler.png')
    pix = ruler_image.load()

    ruler = []
    for i in range(ruler_image.size[1] - 1, -1, -1):
        ruler.append(pix[4, i])

    ruler_image.close()

    return ruler


def parse_config(config: str) -> Dict:
    with open(config, 'r') as f:
        lines = list(filter(lambda y: re.match('\w+=\d+\.?\d*', y), map(lambda x: x.strip(), f.readlines())))

        config_data = {}
        for line in lines:
            result = line.split('=')
            config_data[result[0]] = result[1]

        return config_data


def calc_image_data(ruler: List[float], image: str, config: str) -> Data:
    config_data = parse_config(config)
    ruler_min = float(config_data['ruler_min'])
    ruler_max = float(config_data['ruler_max'])


    target_im = Image.open(image)
    pix = target_im.load()

    print('Calculating {0} ...'.format(image))

    all_value_array = []
    for i in range(0, target_im.size[0]):
        for j in range(0, target_im.size[1]):
            # Only calculate non-transparent pixels
            if pix[i, j][0] != 0 or pix[i, j][1] != 0 or pix[i, j][2] != 0 or pix[i, j][3] != 0:
                idx = ruler.index(min(ruler, key = lambda x: raw_diff(x, pix[i, j])))
                value = (idx / len(ruler)) * (ruler_max - ruler_min) + ruler_min
                all_value_array.append(value)

    target_im.close()

    data = Data(
        image_name = Path(image).stem,
        point_count = len(all_value_array),
        min_value = np.min(all_value_array),
        max_value = np.max(all_value_array),
        mean = np.mean(all_value_array),
        var = np.var(all_value_array),
        std = np.std(all_value_array),
        data_array = all_value_array,
    )

    data.coefficient_of_variation = data.std / data.mean

    return data


def write_output(data_array: List[Data]):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = '总体数据分析'
    ws1['A1'] = '图片'
    ws1['B1'] = '点数'
    ws1['C1'] = '最小值'
    ws1['D1'] = '最大值'
    ws1['E1'] = '均值'
    ws1['F1'] = '方差'
    ws1['G1'] = '标准差'
    ws1['H1'] = '变异系数'

    ws2 = wb.create_sheet(title = '数据集')

    for idx, data in enumerate(data_array):
        ws1.cell(row = idx + 2, column = 1, value = data.image_name)
        ws1.cell(row = idx + 2, column = 2, value = data.point_count)
        ws1.cell(row = idx + 2, column = 3, value = data.min_value)
        ws1.cell(row = idx + 2, column = 4, value = data.max_value)
        ws1.cell(row = idx + 2, column = 5, value = data.mean)
        ws1.cell(row = idx + 2, column = 6, value = data.var)
        ws1.cell(row = idx + 2, column = 7, value = data.std)
        ws1.cell(row = idx + 2, column = 8, value = data.coefficient_of_variation)

        ws2.cell(row = 1, column = idx + 1, value = data.image_name) 
        for data_idx, value in enumerate(data.data_array):
            ws2.cell(row = data_idx + 2, column = idx + 1, value = value)

    wb.save(filename = xlsx_file)


if __name__ == '__main__':
    # Find current script absolute path
    root_path = os.path.dirname(os.path.realpath(__file__))

    # Load ruler
    ruler = load_ruler(root_path)

    # Find images
    images = list(filter(lambda x: x.endswith('.png'), os.listdir('{0}/images'.format(root_path))))
    images.sort()

    # Process each image
    data_array: List[Data] = []
    for image in images:
        image_path = '{0}/images/{1}'.format(root_path, image)
        config_path = '{0}/images/{1}.txt'.format(root_path, os.path.splitext(image)[0])

        if not os.path.isfile(config_path):
            print('Warning: cannot find {0}'.format(config_path))
            continue

        data_array.append(calc_image_data(ruler, image_path, config_path))

    # Write xlsx
    xlsx_file = '{0}/output/output.xlsx'.format(root_path)
    print('Generate {0} ...'.format(xlsx_file))

    if len(data_array) > 0:
        write_output(data_array) 

    print('Done!')
